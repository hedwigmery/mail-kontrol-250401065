import tkinter as tk
from tkinter import messagebox, filedialog
import re
import json
from openpyxl import load_workbook

# Mail kontrol fonksiyonu
def mail_gecerli_mi(email):
    desen = r"^[\w\.-]+@[\w\.-]+\.\w+$"
    return re.match(desen, email) is not None

# Tek mail kontrol
def tek_mail_kontrol():
    email = email_giris.get().strip()

    if email == "":
        messagebox.showwarning("Uyari", "Lutfen e-posta giriniz.")
        return

    if mail_gecerli_mi(email):
        durum = "Gecerli"
        messagebox.showinfo("Sonuc", f"{email}\n\nGecerli e-posta adresi.")
    else:
        durum = "Gecersiz"
        messagebox.showerror("Sonuc", f"{email}\n\nGecersiz e-posta adresi.")

    sonuc = {
        "email": email,
        "durum": durum
    }

    with open("sonuc_tek_mail.json", "w", encoding="utf-8") as dosya:
        json.dump(sonuc, dosya, ensure_ascii=False, indent=4)

# Excel kontrol
def excel_kontrol():
    dosya_yolu = filedialog.askopenfilename(
        title="Excel dosyasi sec",
        filetypes=[("Excel Dosyalari", "*.xlsx")]
    )

    if dosya_yolu == "":
        return

    workbook = load_workbook(dosya_yolu)
    sayfa = workbook.active

    gecerli_sayisi = 0
    gecersiz_sayisi = 0
    sonuclar = []

    for satir in sayfa.iter_rows(values_only=True):
        for hucre in satir:

            if hucre is not None:
                email = str(hucre).strip()

                if "@" in email:

                    if mail_gecerli_mi(email):
                        durum = "Gecerli"
                        gecerli_sayisi += 1
                    else:
                        durum = "Gecersiz"
                        gecersiz_sayisi += 1

                    sonuclar.append({
                        "email": email,
                        "durum": durum
                    })

    with open("sonuc_excel.json", "w", encoding="utf-8") as dosya:
        json.dump(sonuclar, dosya, ensure_ascii=False, indent=4)

    messagebox.showinfo(
        "Excel Sonucu",
        f"Gecerli mail: {gecerli_sayisi}\nGecersiz mail: {gecersiz_sayisi}"
    )

# Pencere
pencere = tk.Tk()
pencere.title("E-posta Kontrol Programi")
pencere.geometry("450x300")

# Baslik
baslik = tk.Label(
    pencere,
    text="E-posta Kontrol Sistemi",
    font=("Arial", 16)
)
baslik.pack(pady=15)

# Tek mail alani
etiket = tk.Label(
    pencere,
    text="Tek e-posta kontrolu:"
)
etiket.pack()

email_giris = tk.Entry(
    pencere,
    width=40
)
email_giris.pack(pady=5)

# Tek mail butonu
tek_buton = tk.Button(
    pencere,
    text="Tek Mail Kontrol Et",
    command=tek_mail_kontrol
)
tek_buton.pack(pady=10)

# Excel alanı
excel_yazi = tk.Label(
    pencere,
    text="Toplu kontrol icin Excel sec:"
)
excel_yazi.pack(pady=10)

excel_buton = tk.Button(
    pencere,
    text="Excel Dosyasi Sec",
    command=excel_kontrol
)
excel_buton.pack(pady=5)

# Programi acik tutar
pencere.mainloop()